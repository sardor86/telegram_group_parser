from aiogram.dispatcher import Dispatcher
from aiogram.types import CallbackQuery
from openpyxl import Workbook

from models import Users, Groups


async def get_all_user(callback: CallbackQuery) -> None:
    wb = Workbook()
    ws = wb.active
    column_sequence = 0
    for group in await Groups().get_all_group():
        ws.merge_cells(start_row=1, start_column=column_sequence * 2 + 1, end_row=1, end_column=column_sequence * 2 + 2)
        ws.cell(row=1, column=column_sequence * 2 + 1).value = group.group_link
        row = 2
        for user in await Users().get_all_user(group.group_link):
            print(user.user_id)
            ws.cell(row=row, column=column_sequence*2+1).value = user.user_id
            ws.cell(row=row, column=column_sequence*2+2).value = user.username
            row += 1
        column_sequence += 1

    wb.save('data.xlsx')
    await callback.bot.send_document(callback.message.chat.id, open('data.xlsx', 'rb'))


def register_get_all_user_handler(dp: Dispatcher) -> None:
    dp.register_callback_query_handler(get_all_user,
                                       lambda callback: callback.data == 'get_all_users')
